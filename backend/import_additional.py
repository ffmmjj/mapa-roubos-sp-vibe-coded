#!/usr/bin/env python3
"""Import ADDITIONAL Excel data into existing database (e.g., 2024 data)."""
import openpyxl
import sqlite3
import os
import sys
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(__file__), "crimes.db")

# Columns to import (same as original)
WANTED_COLS = [
    'ID_DELEGACIA', 'NOME_DEPARTAMENTO', 'NOME_SECCIONAL', 'NOME_DELEGACIA',
    'NOME_MUNICIPIO', 'ANO_BO', 'NUM_BO', 'VERSAO',
    'NOME_DEPARTAMENTO_CIRC', 'NOME_SECCIONAL_CIRC', 'NOME_DELEGACIA_CIRC',
    'NOME_MUNICIPIO_CIRC',
    'DATA_OCORRENCIA_BO', 'HORA_OCORRENCIA', 'DESCRICAO_APRESENTACAO',
    'DATAHORA_REGISTRO_BO', 'DATA_COMUNICACAO_BO',
    'DESCR_PERIODO', 'AUTORIA_BO', 'FLAG_FLAGRANTE', 'FLAG_STATUS',
    'RUBRICA', 'DESCR_CONDUTA', 'DESCR_TIPOLOCAL', 'DESCR_SUBTIPOLOCAL',
    'CIDADE', 'BAIRRO', 'CEP', 'LOGRADOURO', 'NUMERO_LOGRADOURO',
    'LATITUDE', 'LONGITUDE', 'MARCA_OBJETO',
    'FLAG_BLOQUEIO', 'FLAG_DESBLOQUEIO'
]  # Note: MES_REGISTRO_BO and ANO_REGISTRO_BO handled separately


def clean_val(v):
    """Convert None and 'NULL' strings to None for SQLite NULL."""
    if v is None or str(v) == 'NULL':
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if hasattr(v, 'strftime') and not isinstance(v, datetime):
        return v.strftime("%H:%M:%S")
    return v


def import_new_data(xlsx_path: str):
    """Import data from new Excel file into existing database."""
    if not os.path.exists(xlsx_path):
        print(f"❌ Arquivo não encontrado: {xlsx_path}")
        return

    print(f"📂 Arquivo: {xlsx_path}")

    # Check database exists
    if not os.path.exists(DB_PATH):
        print(f"❌ Banco não encontrado: {DB_PATH}")
        print("   Execute import_full.py primeiro para criar o banco inicial.")
        return

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    c = conn.cursor()

    # Check current database state
    c.execute("SELECT COUNT(*) FROM crimes")
    current_total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM crimes WHERE latitude IS NOT NULL AND latitude != 0")
    current_with_coords = c.fetchone()[0]
    c.execute("SELECT MIN(ano_registro_bo), MAX(ano_registro_bo) FROM crimes")
    min_year, max_year = c.fetchone()

    print(f"\n📊 Banco atual:")
    print(f"   Total roubos:          {current_total:,}")
    print(f"   Com coordenadas:      {current_with_coords:,}")
    print(f"   Período:              {min_year} a {max_year}")

    # Open Excel file
    print(f"\n📖 Lendo arquivo Excel...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # Find the correct sheet
    available_sheets = wb.sheetnames
    print(f"   Abas disponíveis: {available_sheets}")

    # Try to find data sheet - more flexible matching
    data_sheet = None
    for sheet_name in ['CELULAR_2024', 'CELULAR_2025', 'CELULAR_2023', 'CELULAR', 'DATA']:
        if sheet_name in available_sheets:
            data_sheet = sheet_name
            break

    if not data_sheet:
        print(f"❌ Não encontrei aba com dados. Abas disponíveis: {available_sheets}")
        wb.close()
        conn.close()
        return

    ws = wb[data_sheet]
    print(f"   Usando aba: {data_sheet}")
    print(f"   Linhas totais: {ws.max_row:,}")

    # Read header
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = list(row)
            # Validate required columns
            required_cols = ['RUBRICA', 'LATITUDE', 'LONGITUDE']
            missing = [c for c in required_cols if c not in header]
            if missing:
                print(f"❌ Colunas faltando: {missing}")
                wb.close()
                conn.close()
                return
            break

    # Find column indices - handle both 2024 and 2025 formats
    col_indices = {}
    mes_idx = None
    ano_idx = None
    
    for col_name in WANTED_COLS:
        if col_name in header:
            col_indices[col_name] = header.index(col_name)
        # Handle 2024 format differences
        elif col_name == 'MES_REGISTRO_BO' and 'MES' in header:
            mes_idx = header.index('MES')
        elif col_name == 'ANO_REGISTRO_BO':
            # ANO_BO is different from ANO_REGISTRO
            if 'ANO' in header and header.index('ANO') < 10:
                # ANO near end is ANO_REGISTRO
                ano_idx = header.index('ANO')

    # Check for duplicates by (num_bo, versao)
    existing_bo = set()
    c.execute("SELECT num_bo, versao FROM crimes")
    for row in c.fetchall():
        existing_bo.add((row[0], row[1]))
    print(f"   Registros existentes no banco: {len(existing_bo):,}")

    # Import new records
    print(f"\n📥 Importando roubos...")
    batch = []
    batch_size = 5000
    imported = 0
    skipped_duplicate = 0
    skipped_not_robbery = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue

        # Extract values
        vals = []
        for col_name in WANTED_COLS:
            idx = col_indices.get(col_name)
            vals.append(clean_val(row[idx]) if idx is not None else None)
        
        # Add MES_REGISTRO_BO and ANO_REGISTRO_BO (handle 2024 format: MES, ANO)
        if mes_idx is not None and ano_idx is not None:
            vals.append(clean_val(row[mes_idx]))  # MES_REGISTRO_BO
            vals.append(clean_val(row[ano_idx]))  # ANO_REGISTRO_BO
        else:
            vals.append(None)  # MES_REGISTRO_BO
            vals.append(None)  # ANO_REGISTRO_BO

        # Check if robbery
        rubrica = vals[WANTED_COLS.index('RUBRICA')] if 'RUBRICA' in col_indices else None
        if not rubrica or 'Roubo' not in rubrica:
            skipped_not_robbery += 1
            continue

        # Check for duplicate
        num_bo = vals[WANTED_COLS.index('NUM_BO')] if 'NUM_BO' in col_indices else None
        versao = vals[WANTED_COLS.index('VERSAO')] if 'VERSAO' in col_indices else None
        if (num_bo, versao) in existing_bo:
            skipped_duplicate += 1
            continue

        # Add to batch
        batch.append(vals)
        imported += 1

        # Insert batch
        if len(batch) >= batch_size:
            placeholders = ",".join(["?"] * (len(WANTED_COLS) + 2))  # +2 for MES_REGISTRO_BO, ANO_REGISTRO_BO
            cols = ",".join([c.lower() for c in WANTED_COLS]) + ",mes_registro_bo,ano_registro_bo"
            c.executemany(f"INSERT INTO crimes ({cols}) VALUES ({placeholders})", batch)
            conn.commit()

            # Add to existing set
            for v in batch:
                num_bo_v = v[WANTED_COLS.index('NUM_BO')] if 'NUM_BO' in WANTED_COLS else None
                versao_v = v[WANTED_COLS.index('VERSAO')] if 'VERSAO' in WANTED_COLS else None
                if num_bo_v and versao_v:
                    existing_bo.add((num_bo_v, versao_v))

            batch = []
            print(f"   Importados: {imported:,}...")

    # Insert remaining
    if batch:
        placeholders = ",".join(["?"] * (len(WANTED_COLS) + 2))  # +2 for MES_REGISTRO_BO, ANO_REGISTRO_BO
        cols = ",".join([c.lower() for c in WANTED_COLS]) + ",mes_registro_bo,ano_registro_bo"
        c.executemany(f"INSERT INTO crimes ({cols}) VALUES ({placeholders})", batch)
        conn.commit()
        print(f"   Importados: {imported:,}...")

    wb.close()

    # Final stats
    c.execute("SELECT COUNT(*) FROM crimes")
    new_total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM crimes WHERE latitude IS NOT NULL AND latitude != 0")
    new_with_coords = c.fetchone()[0]
    c.execute("SELECT MIN(ano_registro_bo), MAX(ano_registro_bo) FROM crimes")
    new_min_year, new_max_year = c.fetchone()

    conn.close()

    print(f"\n✅ Importação concluída!")
    print(f"\n📊 Comparação:")
    print(f"                     Antes    Depois    Diferença")
    print(f"   Total roubos:     {current_total:>7,}  {new_total:>7,}  +{new_total-current_total:>7,}")
    print(f"   Com coordenadas:  {current_with_coords:>7,}  {new_with_coords:>7,}  +{new_with_coords-current_with_coords:>7,}")
    print(f"   Período:          {min_year}-{max_year}       {new_min_year}-{new_max_year}")
    print(f"\n📋 Detalhes da importação:")
    print(f"   Registros novos importados:  {imported:,}")
    print(f"   Pulados (duplicados):       {skipped_duplicate:,}")
    print(f"   Pulados (não roubo):        {skipped_not_robbery:,}")

    if imported > 0:
        print(f"\n⚠️  DICA: Execute o geocoding para resolver coords dos novos registros:")
        print(f"   cd backend")
        print(f"   python3 cross_reference_coords.py")
        print(f"   export MAPBOX_ACCESS_TOKEN='seu_token'")
        print(f"   python3 geocode_mapbox.py")

    print(f"\n🔄 Reinicie o servidor para ver os novos dados:")
    print(f"   lsof -i :8000 -t | xargs kill 2>/dev/null")
    print(f"   ./start.sh")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 import_additional.py <caminho/arquivo.xlsx>")
        print("Exemplo: python3 import_additional.py ../CelularesSubtraidos_2024.xlsx")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    import_new_data(xlsx_path)
